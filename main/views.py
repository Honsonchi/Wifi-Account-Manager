from typing import Any, Dict
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.urls import path, reverse_lazy
from django.shortcuts import render, redirect
from django.contrib.auth.views import PasswordChangeView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views.generic import ListView, UpdateView, DeleteView, CreateView, FormView
from .models import UserInfo, Device, Group, User
from django.contrib.auth.models import Group as groups
from .form import DeviceForm, AdminDeviceForm, BaseUserCreateFormSet, UserCreateForm, UserInfoForm, UploadFileForm
import openpyxl

# 首頁
class HomePage(ListView):
    model=Device

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
            now_user_group = Group.objects.filter(UserData__UserData=self.request.user)
            if now_user_group.count() > 0:
                context['now_user_group'] = now_user_group[0].Name
            else:
                context['now_user_group'] = '無'
        else:
            context['now_user'] = ''
            context['now_user_group'] = '無'
        
        return context
    
    def get_queryset(self):
        if self.request.user.is_authenticated:
            return Device.objects.filter(Owner__UserData=self.request.user)
        return super().get_queryset()
    
    context_object_name = 'devices'
    template_name = 'homepage.html'

# 變更密碼
class PasswordChange(PermissionRequiredMixin, LoginRequiredMixin, PasswordChangeView):
    permission_required = ['main.can_assess']

    login_url = reverse_lazy('login')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
        else:
            context['now_user'] = ''
        return context

    redirect_field_name = reverse_lazy('homepage')
    template_name = 'registration/password.html'

# 裝置管理
class DeviceManaging(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = Device
    permission_required = ['main.can_assess']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
        else:
            context['now_user'] = ''
        return context

    def get_queryset(self):
        return Device.objects.filter(Owner=UserInfo.objects.get(UserData=self.request.user))

    context_object_name = 'devices'
    template_name = 'device_managing.html'

# 裝置編輯
class DeviceEdit(PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    model = Device
    form_class = DeviceForm
    permission_required = ['main.can_assess']
    
    def get_success_url(self):
        return reverse_lazy('device_managing')

    def form_valid(self, form):
        now_user = UserInfo.objects.get(UserData=self.request.user)
        if not self.request.user.has_perm('main.admin'): #如果不是管理員
            if form.instance.Owner == now_user:
                return super().form_valid(form)
            else:
                raise PermissionDenied()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
        else:
            context['now_user'] = ''
        return context

    pk_url_kwarg = 'deviceid'
    context_object_name = 'device'
    template_name = 'device_edit.html'

# 裝置刪除
class DeviceDelete(PermissionRequiredMixin, LoginRequiredMixin, DeleteView):
    model = Device
    permission_required = ['main.can_assess']
    
    def get_success_url(self):
        return reverse_lazy('device_managing')
    
    def form_valid(self, form):
        now_user = UserInfo.objects.get(UserData=self.request.user)
        if not self.request.user.has_perm('main.admin'): #如果不是管理員
            if form.instance.Owner == now_user:
                return super().form_valid(form)
            else:
                raise PermissionDenied()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
        else:
            context['now_user'] = ''
        return context

    pk_url_kwarg = 'deviceid'
    context_object_name = 'device'
    template_name = 'device_delete.html'

# 裝置創建
class DeviceCreate(PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    model = Device
    form_class = DeviceForm
    permission_required = ['main.can_assess']

    def get_success_url(self):
        return reverse_lazy('device_managing')
    
    def form_valid(self, form):
        form.instance.Owner = UserInfo.objects.get(UserData=self.request.user)
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
        else:
            context['now_user'] = ''
        return context

    pk_url_kwarg = 'deviceid'
    context_object_name = 'device'
    template_name = 'device_create.html'

#管理員: 群組管理
class GroupManaging(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = Group
    ordering = ['Name']
    paginate_by = 20
    permission_required = ['main.can_assess', 'main.admin']

    def post(self, request, *args, **kwargs):
        selected_ids = request.POST.getlist('selected_objects')
        if len(selected_ids) > 0:
            for i in Group.objects.filter(id__in=selected_ids):
                for user in i.UserData.all():
                    user.UserData.delete()
            Group.objects.filter(id__in=selected_ids).delete()
            return redirect('.')
        return redirect('.')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
        else:
            context['now_user'] = ''
        return context
    
    context_object_name = 'groups'
    template_name = 'group_managing.html'

#管理員: 群組編輯
class GroupEdit(PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    model = Group
    permission_required = ['main.can_assess', 'main.admin']
    
    fields = ['Name', 'UserData', 'Internet', 'Note']

    def get_success_url(self):
        return reverse_lazy('group_managing')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
        else:
            context['now_user'] = ''
        return context
    
    def form_valid(self, form):
        for i in Group.objects.get(id=form.instance.id).UserData.all():
            i.Internet = form.instance.Internet
            i.save()
        return super().form_valid(form)

    pk_url_kwarg = 'groupid'
    context_object_name = 'group'
    template_name = 'group_edit.html'

# 管理員: 群組刪除
# class GroupDelete(PermissionRequiredMixin, LoginRequiredMixin, DeleteView):
#     model = Group
#     permission_required = ['main.can_assess', 'main.admin']

#     def get_success_url(self):
#         return reverse_lazy('group_managing')

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         if self.request.user.is_authenticated:
#             context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
#         else:
#             context['now_user'] = ''

#         return context

#     pk_url_kwarg = 'groupid'
#     context_object_name = 'group'
#     template_name = 'group_delete.html'

# 裝置創建
class GroupCreate(PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    model = Group
    permission_required = ['main.can_assess', 'main.admin']

    fields = ['Name', 'UserData', 'Internet', 'Note']

    def get_success_url(self):
        return reverse_lazy('group_managing')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
        else:
            context['now_user'] = ''
        return context

    pk_url_kwarg = 'groupid'
    context_object_name = 'group'
    template_name = 'group_create.html'

#管理員管理
class Manage(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = Device
    ordering = ['Owner__UserType', 'Owner__Grade', 'Owner__Class', 'Owner__SeatNumber', 'Name']
    paginate_by = 20
    permission_required = ['main.can_assess', 'main.admin']

    def post(self, request, *args, **kwargs):
        selected_ids = request.POST.getlist('selected_objects')
        if len(selected_ids) > 0:
            Device.objects.filter(id__in=selected_ids).delete()
            return redirect('.')
        return redirect('.')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
        else:
            context['now_user'] = ''
        return context
    
    context_object_name = 'devices'
    template_name = 'manage.html'

#管理員編輯裝置
class AdminDeviceEdit(PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    model = Device
    form_class = AdminDeviceForm
    permission_required = ['main.can_assess', 'main.admin']
    
    def get_success_url(self):
        return reverse_lazy('manage')

    def form_valid(self, form):
        now_user = UserInfo.objects.get(UserData=self.request.user)
        if not self.request.user.has_perm('main.admin'): #如果不是管理員
            if form.instance.Owner == now_user:
                return super().form_valid(form)
            else:
                raise PermissionDenied()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
        else:
            context['now_user'] = ''
        return context

    pk_url_kwarg = 'deviceid'
    context_object_name = 'device'
    template_name = 'admin_device_edit.html'

# 管理員裝置刪除
class AdminDeviceDelete(PermissionRequiredMixin, LoginRequiredMixin, DeleteView):
    model = Device
    permission_required = ['main.can_assess', 'main.admin']
    
    def get_success_url(self):
        return reverse_lazy('manage')
    
    def form_valid(self, form):
        now_user = UserInfo.objects.get(UserData=self.request.user)
        if not self.request.user.has_perm('main.admin'): #如果不是管理員
            if form.instance.Owner == now_user:
                return super().form_valid(form)
            else:
                raise PermissionDenied()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
        else:
            context['now_user'] = ''
        return context

    pk_url_kwarg = 'deviceid'
    context_object_name = 'device'
    template_name = 'admin_device_delete.html'

# 管理員裝置新增
class AdminDeviceCreate(PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    model = Device
    form_class = AdminDeviceForm
    permission_required = ['main.can_assess', 'main.admin']

    def get_success_url(self):
        return reverse_lazy('manage')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
        else:
            context['now_user'] = ''
        return context

    pk_url_kwarg = 'deviceid'
    context_object_name = 'device'
    template_name = 'admin_device_create.html'

#管理員人員管理
class UserManaging(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = UserInfo
    ordering = ['UserType', 'Grade', 'Class', 'SeatNumber', 'Name']
    paginate_by = 20
    permission_required = ['main.can_assess', 'main.admin']

    def post(self, request, *args, **kwargs):
        selected_ids = request.POST.getlist('selected_objects')
        if len(selected_ids) > 0:
            User.objects.filter(id__in=selected_ids).delete()
            return redirect('.')
        return redirect('.')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
        else:
            context['now_user'] = ''
        context['Group'] = Group.objects.all()
        context['Device'] = Device.objects.all()
        return context
    
    context_object_name = 'members'
    template_name = 'user_managing.html'

#管理員新增人員
class UserCreate(PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    model = User
    form_class = UserCreateForm
    permission_required = ['main.can_assess', 'main.admin']

    def get_success_url(self):
        return reverse_lazy('user_managing')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["formset"] = BaseUserCreateFormSet(self.request.POST)
        else:
            context["formset"] = BaseUserCreateFormSet

        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
        else:
            context['now_user'] = ''

        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        if form.is_valid() and formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()

            usertype = formset.cleaned_data[0].get('UserType')
            user_obj = User.objects.get(id=formset.instance.id)
            if usertype == 0:
                user_obj.groups.add(groups.objects.get(name='管理員'))
            elif usertype == 1:
                user_obj.groups.add(groups.objects.get(name='老師'))
            elif usertype == 2:
                user_obj.groups.add(groups.objects.get(name='學生'))

            Group.objects.get(id=formset.cleaned_data[0].get('user_group')).UserData.add(UserInfo.objects.get(UserData=user_obj))

            return super().form_valid(form)
        else:
            return self.render_to_response(self.get_context_data(form=form))

    template_name = 'user_create.html'

#管理員編輯人員
class UserEdit(PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    model = UserInfo
    form_class = UserInfoForm
    # fields = ['Name', 'UserType', 'StuId', 'Email', 'Grade', 'Class', 'SeatNumber', 'Internet']
    permission_required = ['main.can_assess', 'main.admin']

    def get_success_url(self):
        return reverse_lazy('user_managing')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
            context['now_edit_user'] = self.get_object().UserData
        else:
            context['now_user'] = ''
            context['now_edit_user'] = ''

        return context
    
    def get_initial(self):
        initial = super().get_initial()
        initial['user_group'] = Group.objects.get(UserData=self.get_object()).id
        return initial

    def form_valid(self, form):
        user_obj = self.get_object()
        new_group = form.cleaned_data['user_group']
        print(type(new_group))
        print(new_group)
        for g in Group.objects.filter(UserData=user_obj):
            g.UserData.remove(user_obj)
        Group.objects.get(id=new_group).UserData.add(user_obj)

        return super().form_valid(form)

    pk_url_kwarg = 'userid'
    context_object_name = 'now_edit_member'
    template_name = 'user_edit.html'

#管理員人員刪除

class UserDelete(PermissionRequiredMixin, LoginRequiredMixin, DeleteView):
    model = User
    permission_required = ['main.can_assess', 'main.admin']
    
    def get_success_url(self):
        return reverse_lazy('user_managing')
    
    def form_valid(self, form):
        now_user = UserInfo.objects.get(UserData=self.request.user)
        if not self.request.user.has_perm('main.admin'): #如果不是管理員
            if form.instance.Owner == now_user:
                return super().form_valid(form)
            else:
                raise PermissionDenied()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context['now_user'] = UserInfo.objects.get(UserData=self.request.user)
            context['now_delete_member'] = UserInfo.objects.get(UserData=self.get_object())
            context['now_delete_member_device'] = Device.objects.filter(Owner=UserInfo.objects.get(UserData=self.get_object())).count()
        else:
            context['now_user'] = ''
            context['now_delete_member'] = ''
            context['now_delete_member_device'] = 0
        return context

    pk_url_kwarg = 'userid'
    context_object_name = 'now_delete_user'
    template_name = 'user_delete.html'

class BatchCreateUser(PermissionRequiredMixin, LoginRequiredMixin, FormView):
    form_class = UploadFileForm
    template_name = 'batch_create_user.html'
    permission_required = ['main.can_assess', 'main.admin']
    

    success_url = reverse_lazy('user_managing')

    def form_valid(self, form):
        file = form.cleaned_data['file']

        excel_file = openpyxl.load_workbook(file.open())
        wb = excel_file.worksheets[0]

        for row in wb.iter_rows():
            data = list(map(lambda r: r.value, row))
            content_error = False
            for i in range(12):
                if (str(data[i]).startswith('$') and str(data[i]).endswith('$')):
                    content_error = True
                    break
                if data[i] == None:
                    if i == 0:
                        content_error = True
                        break
                    elif i == 6:
                        data[i] = ''
                    elif 5 <= i and i <= 9:
                        data[i] = 0
                    elif i == 10:
                        data[i] = True
                    elif i == 11:
                        data[i] = ''

            if str(data[3]) == '管理員':
                usertype = 0
            elif str(data[3]) == '老師':
                usertype = 1
            elif str(data[3]) == '學生':
                usertype = 2
            else:
                content_error = True

            if content_error:
                continue

            UserObj = User.objects.filter(username=str(data[0]))
            if UserObj.count() > 0:
                UserObj[0].username = str(data[0])
                UserObj[0].set_password(str(data[1]))
                UserObj[0].save()
                UserInfoObj = UserInfo.objects.get(UserData=UserObj[0])
                UserInfoObj.UserType = usertype
                UserInfoObj.Name = str(data[2])
                UserInfoObj.StuId = int(data[5])
                UserInfoObj.Email = str(data[6])
                UserInfoObj.Grade = int(data[7])
                UserInfoObj.Class = int(data[8])
                UserInfoObj.SeatNumber = int(data[9])
                UserInfoObj.Internet = bool(data[10])
                UserInfoObj.Note = str(data[11])
                UserInfoObj.save()
            else:
                for i in range(12):
                    if data[i] == None:
                        if 1 <= i and i <= 4:
                            content_error = True
                            break

                if content_error:
                    continue

                created_user = User.objects.create_user(username=str(data[0]), password=str(data[1]))
                created_user.groups.add(groups.objects.get(name=str(data[3])))
                User.save(created_user)

                created_user_info = UserInfo.objects.create(UserData=created_user,
                                        UserType=usertype,
                                        Name=str(data[2]),
                                        StuId=int(data[5]),
                                        Email=str(data[6]),
                                        Grade=int(data[7]),
                                        Class=int(data[8]),
                                        SeatNumber=int(data[9]),
                                        Internet=bool(data[10]),
                                        Note=str(data[11]))
                UserInfo.save(created_user_info)

                group, created = Group.objects.get_or_create(Name=str(data[4]),
                                                             defaults={'Name': str(data[4])})
                group.UserData.add(created_user_info)

                Group.save(group)

        return super().form_valid(form)
    
class BatchAdminDeviceCreate(PermissionRequiredMixin, LoginRequiredMixin, FormView):
    form_class = UploadFileForm
    template_name = 'batch_create_device.html'
    permission_required = ['main.can_assess', 'main.admin']
    
    success_url = reverse_lazy('manage')

    def form_valid(self, form):
        file = form.cleaned_data['file']

        excel_file = openpyxl.load_workbook(file.open())
        wb = excel_file.worksheets[0]

        for row in wb.iter_rows():
            data = list(map(lambda r: r.value, row))
            content_error = False

            for i in range(4):
                if (str(data[i]).startswith('$') and str(data[i]).endswith('$')):
                    content_error = True
                    break
                if data[i] == None:
                    if i == 3:
                        data[i] = ''
                    else:
                        content_error = True
                        break

            if content_error:
                continue
            
            now_user_obj = User.objects.filter(username=str(data[0]))
            if now_user_obj.count() > 0:
                created_device = Device.objects.create(Owner=UserInfo.objects.get(UserData=now_user_obj[0].id),
                                                       Name=str(data[1]),
                                                       MacAddress=data[2],
                                                       Note=str(data[3]))
                Device.save(created_device)
            else:
                continue
        
        return super().form_valid(form)